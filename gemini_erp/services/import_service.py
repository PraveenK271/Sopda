"""Historical import engine (track H).

A generic read -> validate -> report -> import tool for bringing this financial
year's real trading history in from Excel. Two-stage by contract: validation
reads the whole file and writes NOTHING; the user sees every problem before
deciding to import. Per-type import methods (H2-H5) reuse the existing services
(SalesService, PurchaseService, ...) — there is no parallel save path.

All business rules live here, not in the UI (CLAUDE.md).
"""

import logging
from dataclasses import dataclass, field
from datetime import date as date_type
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from decimal import ROUND_HALF_UP, Decimal

from sqlalchemy import false
from database import get_session
from models import Customer, ImportLog, Item, PurchaseInvoice, SalesInvoice, Supplier
from services.customer_service import CustomerService
from services.gst_service import split_gst
from services.opening_balance_service import OpeningBalanceService
from services.purchase_service import PurchaseService
from services.sales_service import SalesService
from services.supplier_service import SupplierService

logger = logging.getLogger(__name__)

DEFAULT_STATE = "Andhra Pradesh"  # B2C default when customer_state is blank
TOTAL_TOLERANCE = Decimal("1")    # rupees of rounding slack on the total cross-check

# The historical cut-off: transactions on/after this date are entered
# individually; everything before is collapsed into opening balances (H2).
CUTOFF_DATE = date_type(2026, 4, 1)
DATE_FORMAT = "%d-%m-%Y"

# Marker written into an extra NOTE column of the template's example row so the
# reader can skip it even if the user forgets to delete it.
EXAMPLE_MARKER = "EXAMPLE - DELETE ME"


@dataclass(frozen=True)
class ColumnSpec:
    name: str
    kind: str = "text"          # text | date | number
    required: bool = True
    role: str | None = None      # date | item_code | quantity | rate | total |
                                 # amount | opening_qty | party_name | unique_ref |
                                 # party_type | balance_type | mode | bank_account
    description: str = ""
    example: str = ""


@dataclass(frozen=True)
class ImportTypeDef:
    key: str
    party_kind: str | None       # 'CUSTOMER' | 'SUPPLIER' | None
    columns: list[ColumnSpec]

    def column_names(self) -> list[str]:
        return [c.name for c in self.columns]


# --- Column definitions per import type -------------------------------------

IMPORT_DEFS: dict[str, ImportTypeDef] = {
    "SALES": ImportTypeDef("SALES", "CUSTOMER", [
        ColumnSpec("invoice_no", "text", True, "unique_ref", "Bill-book invoice number; must be unique.", "DA001/2026-2027"),
        ColumnSpec("invoice_date", "date", True, "date", "DD-MM-YYYY, on/after 01-04-2026.", "05-04-2026"),
        ColumnSpec("customer_name", "text", True, "party_name", "Auto-created after confirmation if unknown.", "Ramesh Traders"),
        ColumnSpec("customer_gstin", "text", False, None, "Blank for B2C.", "37ABCDE1234F1Z5"),
        ColumnSpec("customer_state", "text", False, None, "Blank defaults to Andhra Pradesh.", "Andhra Pradesh"),
        ColumnSpec("item_code", "text", True, "item_code", "Must already exist in the Item Master.", "ITM-001"),
        ColumnSpec("quantity", "number", True, "quantity", "Greater than 0.", "10"),
        ColumnSpec("rate", "number", True, "rate", "Per-unit rate, excluding GST.", "100"),
        ColumnSpec("invoice_total", "number", True, "total", "Bill-book total (cross-check); repeat on every row of the invoice.", "1180"),
    ]),
    "PURCHASES": ImportTypeDef("PURCHASES", "SUPPLIER", [
        ColumnSpec("bill_no", "text", True, "unique_ref", "Supplier's bill number; unique per supplier.", "S-100"),
        ColumnSpec("bill_date", "date", True, "date", "DD-MM-YYYY, on/after 01-04-2026.", "05-04-2026"),
        ColumnSpec("supplier_name", "text", True, "party_name", "Auto-created after confirmation if unknown.", "Sri Metals"),
        ColumnSpec("supplier_gstin", "text", False, None, "Blank if unregistered.", "29ABCDE1234F1Z5"),
        ColumnSpec("supplier_state", "text", False, None, "Blank defaults to Andhra Pradesh.", "Karnataka"),
        ColumnSpec("item_code", "text", True, "item_code", "Must already exist in the Item Master.", "ITM-001"),
        ColumnSpec("quantity", "number", True, "quantity", "Greater than 0.", "50"),
        ColumnSpec("rate", "number", True, "rate", "Per-unit rate, excluding GST.", "80"),
        ColumnSpec("bill_total", "number", True, "total", "Bill total (cross-check); repeat on every row of the bill.", "4720"),
    ]),
    "OPENING_STOCK": ImportTypeDef("OPENING_STOCK", None, [
        ColumnSpec("item_code", "text", True, "item_code", "Must already exist in the Item Master.", "ITM-001"),
        ColumnSpec("opening_qty", "number", True, "opening_qty", "Stock quantity ON 31-03-2026 (NOT today's count).", "60"),
    ]),
    "OPENING_BALANCES": ImportTypeDef("OPENING_BALANCES", None, [
        ColumnSpec("party_type", "text", True, "party_type", "CUSTOMER or SUPPLIER.", "CUSTOMER"),
        ColumnSpec("party_name", "text", True, "party_name", "Must already exist.", "Ramesh Traders"),
        ColumnSpec("amount", "number", True, "amount", "Opening balance amount (positive).", "50000"),
        ColumnSpec("balance_type", "text", True, "balance_type", "Dr (customer owes us) or Cr (we owe supplier).", "Dr"),
    ]),
    "RECEIPTS": ImportTypeDef("RECEIPTS", "CUSTOMER", [
        ColumnSpec("date", "date", True, "date", "DD-MM-YYYY, on/after 01-04-2026.", "10-04-2026"),
        ColumnSpec("customer_name", "text", True, "party_name", "Must already exist.", "Ramesh Traders"),
        ColumnSpec("amount", "number", True, "amount", "Amount received (positive).", "1180"),
        ColumnSpec("mode", "text", True, "mode", "CASH or BANK.", "BANK"),
        ColumnSpec("bank_account_name", "text", False, "bank_account", "Required when mode is BANK; must already exist.", "HDFC Current"),
        ColumnSpec("reference_no", "text", False, None, "Cheque/UPI reference.", "UPI-12345"),
    ]),
    "PAYMENTS": ImportTypeDef("PAYMENTS", "SUPPLIER", [
        ColumnSpec("date", "date", True, "date", "DD-MM-YYYY, on/after 01-04-2026.", "10-04-2026"),
        ColumnSpec("supplier_name", "text", True, "party_name", "Must already exist.", "Sri Metals"),
        ColumnSpec("amount", "number", True, "amount", "Amount paid (positive).", "4720"),
        ColumnSpec("mode", "text", True, "mode", "CASH or BANK.", "BANK"),
        ColumnSpec("bank_account_name", "text", False, "bank_account", "Required when mode is BANK; must already exist.", "HDFC Current"),
        ColumnSpec("reference_no", "text", False, None, "Cheque/UPI reference.", "CHQ-0091"),
    ]),
}


@dataclass
class ValidationReport:
    """Outcome of validating a file. Holds no DB handle — safe to pass to the UI."""

    import_type: str
    rows_read: int = 0
    errors: list = field(default_factory=list)     # {row_number, message} — block import
    warnings: list = field(default_factory=list)    # {row_number, message} — allow, but show

    def add_error(self, row_number: int, message: str) -> None:
        self.errors.append({"row_number": row_number, "message": message})

    def add_warning(self, row_number: int, message: str) -> None:
        self.warnings.append({"row_number": row_number, "message": message})

    @property
    def is_importable(self) -> bool:
        return len(self.errors) == 0

    @property
    def summary(self) -> dict:
        return {
            "import_type": self.import_type,
            "rows_read": self.rows_read,
            "error_count": len(self.errors),
            "warning_count": len(self.warnings),
            "importable": self.is_importable,
        }


class ImportService:
    """Generic import engine: template -> read -> validate -> (import per type)."""

    # --- reading ---------------------------------------------------------

    def read_sheet(self, file_path: str, expected_columns: list[str]) -> list[dict]:
        """Read the first worksheet into a list of {column_name: value} dicts.

        The header row must contain every expected column (order-independent,
        case-insensitive). A missing/misspelled column is a hard error naming it.
        Blank rows and the template's EXAMPLE row are skipped.
        """
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = next(rows_iter)
        except StopIteration:
            raise ValueError("The file is empty.")

        header = [str(h).strip() if h is not None else "" for h in header]
        header_lower = [h.lower() for h in header]
        expected_lower = {c.lower(): c for c in expected_columns}

        missing = [c for c in expected_columns if c.lower() not in header_lower]
        if missing:
            raise ValueError(
                "Missing or misspelled column(s): "
                + ", ".join(missing)
                + ". Expected columns: "
                + ", ".join(expected_columns)
            )

        index_of = {expected_lower[h]: j for j, h in enumerate(header_lower) if h in expected_lower}

        rows: list[dict] = []
        for raw in rows_iter:
            if raw is None or all(v is None or str(v).strip() == "" for v in raw):
                continue  # blank row
            if any(v is not None and EXAMPLE_MARKER in str(v) for v in raw):
                continue  # leftover example row
            rows.append({name: raw[j] if j < len(raw) else None for name, j in index_of.items()})
        return rows

    # --- validation (writes NOTHING) ------------------------------------

    def validate(self, rows: list[dict], import_type: str) -> ValidationReport:
        """Validate parsed rows against the shared rules. Read-only."""
        defn = IMPORT_DEFS[import_type]
        report = ValidationReport(import_type=import_type, rows_read=len(rows))

        session = get_session()
        try:
            checked_refs: set = set()
            for i, row in enumerate(rows):
                row_no = i + 2  # header is row 1; data begins at row 2
                for col in defn.columns:
                    raw = row.get(col.name)
                    blank = raw is None or str(raw).strip() == ""
                    if blank:
                        if col.required:
                            report.add_error(row_no, f"'{col.name}' is required")
                        continue
                    self._validate_cell(session, report, row_no, col, raw, defn, checked_refs)
            self._validate_type_specific(session, report, rows, defn)
        finally:
            session.close()
        return report

    def _validate_type_specific(self, session, report, rows, defn):
        """Rules that depend on more than one cell / on the import type."""
        if defn.key == "OPENING_BALANCES":
            # The party must already exist (we cannot open a balance on a party
            # that has no ledger account).
            for i, row in enumerate(rows):
                row_no = i + 2
                ptype = str(row.get("party_type") or "").strip().upper()
                pname = str(row.get("party_name") or "").strip()
                if not pname or ptype not in ("CUSTOMER", "SUPPLIER"):
                    continue  # already flagged by the per-cell validators
                model = Customer if ptype == "CUSTOMER" else Supplier
                found = (
                    session.query(model.id)
                    .filter(model.name == pname, model.is_deleted == false())
                    .first()
                )
                if found is None:
                    report.add_error(row_no, f"{ptype.capitalize()} '{pname}' does not exist")
        elif defn.key == "SALES":
            self._validate_sales_groups(session, report, rows)
        elif defn.key == "PURCHASES":
            self._validate_purchase_groups(session, report, rows)

    def _validate_sales_groups(self, session, report, rows):
        """Per-invoice checks: one date + one customer per invoice_no, and the
        computed total (taxable + GST) matches the bill-book invoice_total (±₹1)."""
        groups: dict[str, list] = {}
        for i, row in enumerate(rows):
            inv = str(row.get("invoice_no") or "").strip()
            if inv:
                groups.setdefault(inv, []).append((i + 2, row))

        for inv, entries in groups.items():
            first_row_no = entries[0][0]
            if len({str(r.get("invoice_date") or "").strip() for _, r in entries}) > 1:
                report.add_error(first_row_no, f"Invoice '{inv}' has more than one date")
                continue
            if len({str(r.get("customer_name") or "").strip() for _, r in entries}) > 1:
                report.add_error(first_row_no, f"Invoice '{inv}' has more than one customer")
                continue
            state = self._resolve_party_state(
                session, Customer, entries[0][1].get("customer_name"), entries[0][1].get("customer_state")
            )
            self._check_total(report, entries, "invoice_total", session, state, f"Invoice '{inv}'")

    def _validate_purchase_groups(self, session, report, rows):
        """Per-bill checks. A purchase bill is identified by (supplier, bill_no) —
        bill_no is unique PER SUPPLIER, not globally (two suppliers may both use
        '001'). One date per bill; the computed total matches bill_total; and the
        (supplier, bill_no) pair must not already exist in the database."""
        groups: dict[tuple, list] = {}
        for i, row in enumerate(rows):
            bill = str(row.get("bill_no") or "").strip()
            supplier = str(row.get("supplier_name") or "").strip()
            if bill and supplier:
                groups.setdefault((supplier, bill), []).append((i + 2, row))

        for (supplier, bill), entries in groups.items():
            first_row_no = entries[0][0]
            if len({str(r.get("bill_date") or "").strip() for _, r in entries}) > 1:
                report.add_error(first_row_no, f"Bill '{bill}' for '{supplier}' has more than one date")
                continue
            # Duplicate only if THIS supplier already has a bill with this number.
            existing_supplier = (
                session.query(Supplier)
                .filter(Supplier.name == supplier, Supplier.is_deleted == false())
                .first()
            )
            if existing_supplier is not None:
                dup = (
                    session.query(PurchaseInvoice.id)
                    .filter(
                        PurchaseInvoice.invoice_no == bill,
                        PurchaseInvoice.supplier_id == existing_supplier.id,
                        PurchaseInvoice.is_deleted == false(),
                    )
                    .first()
                )
                if dup is not None:
                    report.add_error(first_row_no, f"Bill '{bill}' already exists for supplier '{supplier}'")
                    continue
            state = self._resolve_party_state(
                session, Supplier, supplier, entries[0][1].get("supplier_state")
            )
            self._check_total(report, entries, "bill_total", session, state, f"Bill '{bill}' for '{supplier}'")

    def _check_total(self, report, entries, total_col, session, state, label):
        """Compare the computed taxable+GST for a group's lines to its sheet total.
        Skips silently if a line has a per-cell error (already reported)."""
        computed = Decimal("0")
        for _, r in entries:
            item = (
                session.query(Item)
                .filter(Item.code == str(r.get("item_code") or "").strip(), Item.is_deleted == false())
                .first()
            )
            qty = self._parse_number(r.get("quantity"))
            rate = self._parse_number(r.get("rate"))
            if item is None or qty is None or rate is None:
                return  # a per-cell error already covers this group
            amount = (Decimal(str(qty)) * Decimal(str(rate))).quantize(Decimal("0.01"), ROUND_HALF_UP)
            cgst, sgst, igst = split_gst(amount, item.gst_rate, state)
            computed += amount + cgst + sgst + igst

        sheet_total = self._parse_number(entries[0][1].get(total_col))
        if sheet_total is None:
            return
        if abs(computed - Decimal(str(sheet_total))) > TOTAL_TOLERANCE:
            report.add_error(
                entries[0][0],
                f"{label} total mismatch: computed ₹{computed:.2f} but sheet says ₹{sheet_total:.2f}",
            )

    @staticmethod
    def _resolve_party_state(session, model, name, sheet_state):
        """The state create_*_invoice will use for GST: an existing party's stored
        state, else the sheet's (default Andhra Pradesh for a to-be-created party)."""
        n = str(name or "").strip()
        existing = session.query(model).filter(model.name == n, model.is_deleted == false()).first()
        if existing is not None:
            return existing.state or DEFAULT_STATE
        return str(sheet_state).strip() if sheet_state and str(sheet_state).strip() else DEFAULT_STATE

    def _validate_cell(self, session, report, row_no, col, raw, defn, checked_refs):
        role = col.role
        if col.kind == "date" or role == "date":
            parsed = self._parse_date(raw)
            if parsed is None:
                report.add_error(row_no, f"'{col.name}' must be a date as DD-MM-YYYY (got '{raw}')")
            elif parsed < CUTOFF_DATE:
                report.add_error(row_no, f"'{col.name}' {raw} is before the cut-off 01-04-2026")
            return

        if col.kind == "number":
            value = self._parse_number(raw)
            if value is None:
                report.add_error(row_no, f"'{col.name}' must be a number (got '{raw}')")
                return
            if role == "quantity" and value <= 0:
                report.add_error(row_no, f"'{col.name}' must be greater than 0")
            elif role in ("rate", "total", "opening_qty") and value < 0:
                report.add_error(row_no, f"'{col.name}' cannot be negative")
            elif role == "amount" and value <= 0:
                report.add_error(row_no, f"'{col.name}' must be greater than 0")
            return

        text = str(raw).strip()
        if role == "item_code":
            exists = (
                session.query(Item.id)
                .filter(Item.code == text, Item.is_deleted == false())
                .first()
            )
            if exists is None:
                report.add_error(row_no, f"Item code '{text}' does not exist in the Item Master")
        elif role == "party_name":
            model = Customer if defn.party_kind == "CUSTOMER" else Supplier
            if defn.party_kind:
                found = (
                    session.query(model.id)
                    .filter(model.name == text, model.is_deleted == false())
                    .first()
                )
                if found is None:
                    kind = defn.party_kind.lower()
                    report.add_warning(row_no, f"{kind.capitalize()} '{text}' is unknown; it will be created on import")
        elif role == "unique_ref":
            if text not in checked_refs:
                checked_refs.add(text)
                if defn.key == "SALES":
                    dup = (
                        session.query(SalesInvoice.id)
                        .filter(SalesInvoice.invoice_no == text, SalesInvoice.is_deleted == false())
                        .first()
                    )
                    if dup is not None:
                        report.add_error(row_no, f"Invoice number '{text}' already exists in the database")
                # PURCHASES bill_no is unique PER SUPPLIER — handled in H4.
        elif role == "party_type":
            if text.upper() not in ("CUSTOMER", "SUPPLIER"):
                report.add_error(row_no, f"party_type must be CUSTOMER or SUPPLIER (got '{text}')")
        elif role == "balance_type":
            if text.capitalize() not in ("Dr", "Cr"):
                report.add_error(row_no, f"balance_type must be Dr or Cr (got '{text}')")
        elif role == "mode":
            if text.upper() not in ("CASH", "BANK"):
                report.add_error(row_no, f"mode must be CASH or BANK (got '{text}')")

    @staticmethod
    def _parse_date(raw):
        """Return a date, or None if it cannot be parsed unambiguously."""
        if isinstance(raw, datetime):
            return raw.date()
        if isinstance(raw, date_type):
            return raw
        try:
            return datetime.strptime(str(raw).strip(), DATE_FORMAT).date()
        except (ValueError, TypeError):
            return None

    @staticmethod
    def _parse_number(raw):
        if isinstance(raw, (int, float)):
            return float(raw)
        try:
            return float(str(raw).strip().replace(",", ""))
        except (ValueError, TypeError):
            return None

    # --- template generation --------------------------------------------

    def generate_template(self, import_type: str, save_path: str) -> str:
        """Write a correctly-headed .xlsx template for the given import type."""
        defn = IMPORT_DEFS[import_type]
        wb = Workbook()
        ws = wb.active
        ws.title = import_type[:31]

        bold = Font(bold=True)
        example_fill = PatternFill("solid", fgColor="FFF3CD")

        # Header row (+ a NOTE column that carries the delete-me marker).
        for j, col in enumerate(defn.columns, start=1):
            cell = ws.cell(row=1, column=j, value=col.name)
            cell.font = bold
            ws.column_dimensions[get_column_letter(j)].width = max(14, len(col.name) + 2)
        note_col = len(defn.columns) + 1
        note_cell = ws.cell(row=1, column=note_col, value="NOTE")
        note_cell.font = bold
        ws.column_dimensions[get_column_letter(note_col)].width = 22

        # Force date columns to Text so Excel cannot silently reformat them.
        for j, col in enumerate(defn.columns, start=1):
            if col.kind == "date":
                for r in range(2, 500):
                    ws.cell(row=r, column=j).number_format = "@"

        # One example row, clearly marked.
        for j, col in enumerate(defn.columns, start=1):
            c = ws.cell(row=2, column=j, value=col.example)
            c.fill = example_fill
        marker = ws.cell(row=2, column=note_col, value=EXAMPLE_MARKER)
        marker.fill = example_fill
        marker.font = Font(bold=True, color="9C6500")

        ws.freeze_panes = "A2"

        # Instructions sheet.
        info = wb.create_sheet("Instructions")
        info.cell(row=1, column=1, value=f"{import_type} import — column guide").font = bold
        info.cell(row=2, column=1, value="Delete the yellow EXAMPLE row before importing. Dates are DD-MM-YYYY.")
        headers = ["Column", "Required", "Description"]
        for j, h in enumerate(headers, start=1):
            info.cell(row=4, column=j, value=h).font = bold
        for i, col in enumerate(defn.columns, start=5):
            info.cell(row=i, column=1, value=col.name)
            info.cell(row=i, column=2, value="Yes" if col.required else "No")
            info.cell(row=i, column=3, value=col.description)
        for j, width in enumerate((20, 10, 70), start=1):
            info.column_dimensions[get_column_letter(j)].width = width

        wb.save(save_path)
        logger.info("Generated %s template at %s", import_type, save_path)
        return save_path

    # --- import dispatch -------------------------------------------------

    def import_data(self, import_type: str, file_path: str, created_by: str | None) -> ImportLog:
        """Dispatch to the per-type importer (each re-validates first)."""
        if import_type == "OPENING_STOCK":
            return self.import_opening_stock(file_path, created_by)
        if import_type == "OPENING_BALANCES":
            return self.import_opening_balances(file_path, created_by)
        if import_type == "SALES":
            return self.import_sales(file_path, created_by)
        if import_type == "PURCHASES":
            return self.import_purchases(file_path, created_by)
        raise NotImplementedError(
            f"{import_type} import is added in a later milestone (H5). "
            "Validation and templates are available now."
        )

    def _guard_importable(self, file_path: str, import_type: str) -> list[dict]:
        """Re-read + re-validate; refuse to import a file with any error."""
        rows = self.read_sheet(file_path, IMPORT_DEFS[import_type].column_names())
        report = self.validate(rows, import_type)
        if not report.is_importable:
            first = report.errors[0]
            raise ValueError(
                f"File has {len(report.errors)} error(s); cannot import. "
                f"First: row {first['row_number']} — {first['message']}"
            )
        return rows

    def import_opening_stock(self, file_path: str, created_by: str | None) -> ImportLog:
        rows = self._guard_importable(file_path, "OPENING_STOCK")
        obs = OpeningBalanceService()
        session = get_session()
        created = 0
        try:
            for row in rows:
                item = (
                    session.query(Item)
                    .filter(Item.code == str(row["item_code"]).strip(), Item.is_deleted == false())
                    .first()
                )
                obs.set_opening_stock(session, item.id, self._parse_number(row["opening_qty"]), created_by)
                created += 1
            log = self._write_log(session, "OPENING_STOCK", file_path, len(rows), created, "IMPORTED")
            return log
        except Exception as exc:
            self._write_log(session, "OPENING_STOCK", file_path, len(rows), created, "FAILED", str(exc))
            raise
        finally:
            session.close()

    def import_opening_balances(self, file_path: str, created_by: str | None) -> ImportLog:
        rows = self._guard_importable(file_path, "OPENING_BALANCES")
        obs = OpeningBalanceService()
        session = get_session()
        created = 0
        try:
            for row in rows:
                ptype = str(row["party_type"]).strip().upper()
                pname = str(row["party_name"]).strip()
                model = Customer if ptype == "CUSTOMER" else Supplier
                party = (
                    session.query(model).filter(model.name == pname, model.is_deleted == false()).first()
                )
                obs.set_party_opening_balance(
                    session, ptype, party.id, self._parse_number(row["amount"]),
                    str(row["balance_type"]).strip(), created_by,
                )
                created += 1
            # NOTE: staging only — the user posts the opening journal separately.
            log = self._write_log(session, "OPENING_BALANCES", file_path, len(rows), created, "IMPORTED")
            return log
        except Exception as exc:
            self._write_log(session, "OPENING_BALANCES", file_path, len(rows), created, "FAILED", str(exc))
            raise
        finally:
            session.close()

    def import_sales(self, file_path: str, created_by: str | None) -> ImportLog:
        """Import backdated sales invoices, grouped by invoice_no.

        Each invoice is created via the existing SalesService.create_invoice() —
        GST split, stock OUT rows and the journal entry all come for free — and
        is its own transaction. If one invoice fails, earlier ones stay saved and
        the run stops naming the failed invoice (no whole-file rollback, no silent
        skip). Negative stock is allowed (M28 oversell is "allow but warn"); any
        item that went negative is recorded in the ImportLog notes for H6.
        """
        rows = self._guard_importable(file_path, "SALES")
        sales = SalesService()
        customers = CustomerService()

        # Preserve sheet order; group lines by invoice number.
        grouped: dict[str, list] = {}
        for row in rows:
            grouped.setdefault(str(row["invoice_no"]).strip(), []).append(row)

        created = 0
        stock_notes: list[str] = []
        current = None
        session = get_session()
        try:
            for inv_no, lines in grouped.items():
                current = inv_no
                head = lines[0]
                customer_id = self._resolve_or_create_customer(session, customers, head, created_by)
                invoice_date = self._parse_date(head["invoice_date"])
                line_payload = []
                for r in lines:
                    item = (
                        session.query(Item)
                        .filter(Item.code == str(r["item_code"]).strip(), Item.is_deleted == false())
                        .first()
                    )
                    line_payload.append({
                        "item_id": item.id,
                        "quantity": self._parse_number(r["quantity"]),
                        "rate": self._parse_number(r["rate"]),
                    })
                invoice = sales.create_invoice(
                    invoice_no=inv_no,
                    invoice_date=invoice_date,
                    customer_id=customer_id,
                    lines=line_payload,
                    created_by=created_by,
                )
                created += 1
                for w in getattr(invoice, "stock_warnings", []):
                    stock_notes.append(f"{inv_no}: {w}")

            notes = None
            if stock_notes:
                notes = "Items that went negative during import (review in H6):\n" + "\n".join(stock_notes)
            return self._write_log(session, "SALES", file_path, len(rows), created, "IMPORTED", notes)
        except Exception as exc:
            self._write_log(
                session, "SALES", file_path, len(rows), created, "FAILED",
                f"Stopped at invoice '{current}' after {created} imported: {exc}",
            )
            raise
        finally:
            session.close()

    def import_purchases(self, file_path: str, created_by: str | None) -> ImportLog:
        """Import purchase bills, grouped by (supplier, bill_no), via the existing
        PurchaseService.create_purchase_invoice() — stock IN rows and the
        Dr Purchase / Dr Input GST / Cr Supplier journal come for free. bill_no is
        unique per supplier (not global). Same one-transaction-per-bill,
        stop-on-fail behaviour as sales."""
        rows = self._guard_importable(file_path, "PURCHASES")
        purchases = PurchaseService()
        suppliers = SupplierService()

        grouped: dict[tuple, list] = {}
        for row in rows:
            key = (str(row["supplier_name"]).strip(), str(row["bill_no"]).strip())
            grouped.setdefault(key, []).append(row)

        created = 0
        current = None
        session = get_session()
        try:
            for (supplier_name, bill_no), lines in grouped.items():
                current = f"{bill_no} ({supplier_name})"
                head = lines[0]
                supplier_id = self._resolve_or_create_supplier(session, suppliers, head, created_by)
                bill_date = self._parse_date(head["bill_date"])
                line_payload = []
                for r in lines:
                    item = (
                        session.query(Item)
                        .filter(Item.code == str(r["item_code"]).strip(), Item.is_deleted == false())
                        .first()
                    )
                    line_payload.append({
                        "item_id": item.id,
                        "quantity": self._parse_number(r["quantity"]),
                        "rate": self._parse_number(r["rate"]),
                    })
                purchases.create_purchase_invoice(
                    invoice_no=bill_no,
                    invoice_date=bill_date,
                    supplier_id=supplier_id,
                    lines=line_payload,
                    created_by=created_by,
                )
                created += 1
            return self._write_log(session, "PURCHASES", file_path, len(rows), created, "IMPORTED")
        except Exception as exc:
            self._write_log(
                session, "PURCHASES", file_path, len(rows), created, "FAILED",
                f"Stopped at bill '{current}' after {created} imported: {exc}",
            )
            raise
        finally:
            session.close()

    def _resolve_or_create_supplier(self, session, supplier_service, head, created_by) -> int:
        name = str(head["supplier_name"]).strip()
        existing = (
            session.query(Supplier)
            .filter(Supplier.name == name, Supplier.is_deleted == false())
            .first()
        )
        if existing is not None:
            return existing.id
        gstin = str(head.get("supplier_gstin") or "").strip() or None
        state = str(head.get("supplier_state") or "").strip() or DEFAULT_STATE
        supplier = supplier_service.add_supplier(
            name=name, gstin=gstin, state=state, created_by=created_by
        )
        return supplier.id

    def _resolve_or_create_customer(self, session, customer_service, head, created_by) -> int:
        """Return the customer id, creating the customer if the name is unknown
        (the user confirmed by clicking Import — track Decision 6)."""
        name = str(head["customer_name"]).strip()
        existing = (
            session.query(Customer)
            .filter(Customer.name == name, Customer.is_deleted == false())
            .first()
        )
        if existing is not None:
            return existing.id
        gstin = str(head.get("customer_gstin") or "").strip() or None
        state = str(head.get("customer_state") or "").strip() or DEFAULT_STATE
        customer = customer_service.add_customer(
            name=name, gstin=gstin, state=state, created_by=created_by
        )
        return customer.id

    @staticmethod
    def _write_log(session, import_type, file_path, rows_read, created, status, notes=None) -> ImportLog:
        import os

        log = ImportLog(
            file_name=os.path.basename(file_path),
            import_type=import_type,
            rows_read=rows_read,
            records_created=created,
            status=status,
            notes=notes,
        )
        session.add(log)
        session.commit()
        session.refresh(log)
        session.expunge(log)
        return log
